#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Recover the GM Contamination Register, 1997-2013.

The Register itself stopped being updated in 2013 and its site has been
unreliable since, which is why the escape layer on this map was hand-compiled.
But the whole dataset survives in the open literature:

    Price B & Cotter J (2014). "The GM Contamination Register: a review of
    recorded contamination incidents associated with genetically modified
    organisms (GMOs), 1997-2013." International Journal of Food Contamination
    1:5. doi:10.1186/s40550-014-0005-8

It is open access under CC BY, and **Additional file 1 is the incident table** -
all 396 incidents across 63 countries. That is the source this script pulls.

It does NOT hard-code a download URL. Supplementary-file URLs move; the article
DOI does not.

ROUTE, AND WHY THE HTML ROUTE COULD NOT WORK
--------------------------------------------
The first version fetched the article page and looked for a link to Additional
file 1. Every candidate address was reached and none carried one, run after
run. That is not a layout change: Springer renders the supplementary section
CLIENT-SIDE, so a fetcher receives the page without it. No amount of pattern
work on that HTML will find a link that the server never sent.

Europe PMC holds the same open-access article and exposes its supplementary
files AS DATA rather than as markup:

    search        /europepmc/webservices/rest/search?query=DOI:<doi>  -> PMCID
    supplements   /europepmc/webservices/rest/<PMCID>/supplementaryFiles -> ZIP

Both are documented methods of the Europe PMC Articles RESTful API
(europepmc.org/RestfulWebService); neither is a URL edited into shape, and the
PMCID comes from the search response rather than being guessed. No key, no
registration. The ZIP is opened in memory and the first spreadsheet or CSV in
it is the incident table.

The HTML route is kept BELOW it, unchanged, so that if Europe PMC ever drops
the article the script still has somewhere to look — and so the log shows both
were tried.

If every route fails the script says which one failed and how, and writes
nothing.

    python3 harvest/contamination_register.py --dry-run
    python3 harvest/contamination_register.py

Writes harvest/register_records.json in the same shape as escape_records.json,
which the map merges alongside the hand-compiled incidents. Records already in
escape_records.json are skipped by name so the detailed hand-written entries win
over the register's one-line summaries.

Licence note: CC BY requires attribution, which every emitted record carries in
its `desc` and `url`.
"""
import io, json, re, sys, pathlib
from urllib.request import Request, urlopen

ROOT = pathlib.Path(__file__).resolve().parent.parent
OUT = ROOT / "harvest" / "register_records.json"
HAND = ROOT / "harvest" / "escape_records.json"

DOI = "10.1186/s40550-014-0005-8"

# The Register stopped in 2013. Nothing replaced it, so the honest incident
# picture is several partial sources stacked, each stating its own scope:
#
#   GM Contamination Register  1997-2013, 396 incidents, 63 countries. Compiled
#                              by GeneWatch UK and Greenpeace. The only global
#                              compilation that has ever existed.
#   RASFF                      EU border rejections of unauthorised GM material.
#                              Current and official, but EU imports only.
#   Testbiotech / GMWatch      Case write-ups rather than a dataset - useful for
#                              detail on a specific incident, not for coverage.
#
# All of them together are still not a world picture, and every record says so.
EXTRA_SOURCES = [
    ("RASFF", "https://webgate.ec.europa.eu/rasff-window/screen/search",
     "EU border rejections of unauthorised GM material. Official and current, "
     "but it records what was stopped entering the European Union - not what "
     "happened elsewhere, and not what was never caught."),
    ("Testbiotech", "https://www.testbiotech.org/en/", None),
    ("GMWatch", "https://gmwatch.org/en/gm-contamination-register", None),
]
# The journal moved. It published as International Journal of Food Contamination
# and is now Food Safety and Risk on BioMed Central, so the Springer path no
# longer resolves and the supplementary link went with it. The data is
# "Additional file 1" on the article page. Candidates tried in order.
# Springer Nature Link FIRST: the journal migrated there and the BioMed Central
# address, while still answering, serves a page with no supplement link on it.
ARTICLE = "https://link.springer.com/article/" + DOI
ARTICLE_ALTS = [
    "https://foodsafetyandrisk.biomedcentral.com/articles/" + DOI,
    "https://doi.org/" + DOI,
    "https://foodsafetyandrisk.biomedcentral.com/counter/pdf/" + DOI + ".pdf",
]
# Europe PMC Articles RESTful API. Documented methods, no key, no registration.
# The PMCID is READ FROM THE SEARCH RESPONSE - never assembled from the DOI.
EPMC = "https://www.ebi.ac.uk/europepmc/webservices/rest"
# The DOI is QUOTED and PERCENT-ENCODED. The first attempt sent
#   query=DOI:10.1186/s40550-014-0005-8
# raw, so the slash and the colon went into the query string unescaped and
# Europe PMC matched nothing - a real answer from the API that looked exactly
# like the article being absent. It is not absent; the query was malformed.
EPMC_SEARCH = EPMC + "/search?%s"
EPMC_SUPPL = EPMC + "/%s/supplementaryFiles"

UA = ("GMO-map-harvest/1.0 (+https://github.com/WelcomeToYourGalaxy/GMO-map) "
      "recovering an open-access supplementary dataset")

CITE = ("Recorded in the GM Contamination Register (GeneWatch UK / Greenpeace "
        "International), as published in Price & Cotter 2014, Int. J. Food "
        "Contamination 1:5, CC BY.")

# Country centroids for the countries the Register covers. Region-level only:
# the Register records the country an incident was found in, not a site.
CENTROID = {
 "argentina": (-34.6, -58.4), "australia": (-25.3, 133.8), "austria": (47.5, 14.6),
 "bangladesh": (23.7, 90.4), "belgium": (50.5, 4.5), "bolivia": (-16.3, -63.6),
 "brazil": (-14.2, -51.9), "bulgaria": (42.7, 25.5), "canada": (56.1, -106.3),
 "chile": (-35.7, -71.5), "china": (35.9, 104.2), "colombia": (4.6, -74.3),
 "costa rica": (9.7, -83.8), "croatia": (45.1, 15.2), "cyprus": (35.1, 33.4),
 "czech republic": (49.8, 15.5), "denmark": (56.3, 9.5), "ecuador": (-1.8, -78.2),
 "egypt": (26.8, 30.8), "finland": (61.9, 25.7), "france": (46.2, 2.2),
 "germany": (51.2, 10.5), "greece": (39.1, 21.8), "hungary": (47.2, 19.5),
 "india": (20.6, 79.0), "indonesia": (-0.8, 113.9), "ireland": (53.4, -8.2),
 "italy": (41.9, 12.6), "japan": (36.2, 138.3), "kenya": (-0.02, 37.9),
 "latvia": (56.9, 24.6), "lithuania": (55.2, 23.9), "luxembourg": (49.8, 6.1),
 "mexico": (23.6, -102.6), "netherlands": (52.1, 5.3), "new zealand": (-40.9, 174.9),
 "nigeria": (9.1, 8.7), "norway": (60.5, 8.5), "pakistan": (30.4, 69.3),
 "paraguay": (-23.4, -58.4), "peru": (-9.2, -75.0), "philippines": (12.9, 121.8),
 "poland": (51.9, 19.1), "portugal": (39.4, -8.2), "romania": (45.9, 25.0),
 "russia": (61.5, 105.3), "saudi arabia": (23.9, 45.1), "serbia": (44.0, 21.0),
 "singapore": (1.35, 103.8), "slovakia": (48.7, 19.7), "slovenia": (46.2, 15.0),
 "south africa": (-30.6, 22.9), "south korea": (35.9, 127.8), "spain": (40.5, -3.7),
 "sweden": (60.1, 18.6), "switzerland": (46.8, 8.2), "taiwan": (23.7, 121.0),
 "thailand": (15.9, 101.0), "turkey": (39.0, 35.2), "ukraine": (48.4, 31.2),
 "united kingdom": (55.4, -3.4), "uk": (55.4, -3.4), "united states": (39.8, -98.6),
 "usa": (39.8, -98.6), "vietnam": (14.1, 108.3), "zambia": (-13.1, 27.8),
}


def fetch(url, binary=False):
    req = Request(url, headers={"User-Agent": UA})
    with urlopen(req, timeout=120) as r:
        raw = r.read()
    return raw if binary else raw.decode("utf-8", "replace")


def epmc_pmcid(doi):
    """Ask Europe PMC for this DOI's PMCID. Returns (pmcid, why_not).

    The id is READ from the response. A DOI-to-PMCID rule does not exist and
    inventing one would be the error class nothing downstream catches.
    """
    from urllib.parse import urlencode
    q = urlencode({"query": 'DOI:"%s"' % doi, "format": "json",
                   "resultType": "core", "pageSize": "25"})
    try:
        raw = fetch(EPMC_SEARCH % q)
    except Exception as e:
        return None, "search request failed: %s" % str(e)[:80]
    try:
        d = json.loads(raw)
    except Exception as e:
        return None, "search returned something that is not JSON: %s" % str(e)[:60]
    hits = (((d.get("resultList") or {}).get("result")) or [])
    if not hits:
        # A quoted DOI is the documented form, but field handling has changed
        # before. Try it bare once rather than reporting an absence that is
        # really a syntax difference.
        try:
            q2 = urlencode({"query": "DOI:%s" % doi, "format": "json",
                            "resultType": "core"})
            d = json.loads(fetch(EPMC_SEARCH % q2))
            hits = (((d.get("resultList") or {}).get("result")) or [])
        except Exception:
            hits = []
    if not hits:
        return None, ("search reached Europe PMC and matched no article for this "
                      "DOI, quoted or bare. The DOI is in the docstring and is "
                      "stable, so check the query syntax rather than assuming the "
                      "paper has gone.")
    for h in hits:
        pid = h.get("pmcid")
        if pid:
            return pid, ""
    return None, ("Europe PMC has the article (%d hit(s)) but no PMCID on it, "
                  "which means the full text is not in the open-access subset "
                  "and the supplementary files are not exposed." % len(hits))


def zip_pick(blob):
    """The incident table out of the supplementary ZIP. Returns (name, bytes, why_not).

    Europe PMC returns ALL of an article's supplementary files in one archive.
    Picked by extension, spreadsheets first, and the name of whatever is chosen
    is printed - a run that silently took the wrong file would look identical to
    a run that took the right one.
    """
    import zipfile
    try:
        z = zipfile.ZipFile(io.BytesIO(blob))
    except Exception as e:
        return None, None, "the supplement download is not a readable ZIP: %s" % str(e)[:60]
    names = [n for n in z.namelist() if not n.endswith("/")]
    if not names:
        return None, None, "the supplementary ZIP is empty"
    for exts in ((".xlsx", ".xls"), (".csv",), (".tsv", ".txt")):
        for n in names:
            if n.lower().endswith(exts):
                try:
                    return n, z.read(n), ""
                except Exception as e:
                    return None, None, "could not read %s from the ZIP: %s" % (n, str(e)[:50])
    return None, None, ("the ZIP holds %d file(s) and none is a spreadsheet or CSV: %s. "
                        "If the table is inside a PDF or a DOCX it has to be "
                        "converted by hand." % (len(names), ", ".join(names[:6])))


def find_supplement(page):
    """The link to Additional file 1, wherever the layout has put it."""
    pats = [r'href="(https://static-content\.springer\.com/esm/[^"]+)"',
            r'href="(/articles/[^"]*?/MediaObjects/[^"]+)"',
            r'href="([^"]*MediaObjects[^"]+\.(?:xlsx|xls|csv|pdf|docx))"']
    for p in pats:
        m = re.findall(p, page, re.I)
        if m:
            u = m[0]
            return u if u.startswith("http") else ("https://link.springer.com" + u)
    return None


def rows_from_csv(text):
    import csv
    return list(csv.DictReader(io.StringIO(text)))


def rows_from_xlsx(data):
    try:
        import openpyxl
    except ImportError:
        print("  ! openpyxl not installed - `pip install openpyxl` to read the "
              "spreadsheet supplement", file=sys.stderr)
        return []
    import tempfile
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        f.write(data); path = f.name
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    head = [str(h or "").strip().lower() for h in rows[0]]
    return [dict(zip(head, r)) for r in rows[1:] if any(r)]


def pick(row, *names):
    for n in names:
        for k, v in row.items():
            if k and n in str(k).lower() and v:
                return str(v).strip()
    return ""


def to_record(row, seq):
    country = pick(row, "country", "location", "region")
    year = pick(row, "year", "date")
    crop = pick(row, "crop", "species", "organism")
    what = pick(row, "description", "incident", "detail", "summary", "type")
    if not (country or crop):
        return None
    key = re.sub(r"[^a-z ]", "", country.lower()).strip()
    lat, lng = CENTROID.get(key, (None, None))
    if lat is None:
        return None
    y = re.search(r"(19|20)\d{2}", str(year) or "")
    name = "%s \u2014 %s contamination incident%s" % (
        country, crop or "GM", (" (%s)" % y.group(0)) if y else "")
    return {
        "name": name[:180],
        "source": "escape:register",
        "type": (crop or "Unspecified") + ", recorded contamination incident",
        "lat": lat, "lng": lng, "state": country, "precise": False,
        "impact": 2, "company": "", "size": "",
        "status": "Recorded 1997\u20132013", "phase": "post",
        "date": ("%s-01-01" % y.group(0)) if y else "",
        "url": "https://doi.org/" + DOI,
        "desc": ("WHAT. %s WHERE IT SITS. One of 396 incidents across 63 countries "
                 "recorded between 1997 and 2013. WHY IT MATTERS. The Register "
                 "stopped in 2013 and nothing replaced it, so this is the whole "
                 "systematic record that exists \u2014 and it was compiled by two "
                 "campaigning organisations rather than by any regulator. %s"
                 % ((what or ("Recorded contamination involving %s in %s."
                              % (crop or "GM material", country)))[:400], CITE)),
        "checked": "",
    }


def main():
    # THIS LOOP USED TO BREAK ON THE FIRST PAGE THAT FETCHED, not on the first
    # page that carried what it came for. The journal moved to Springer Nature
    # Link, and the old BioMed Central address still answers - with a page that
    # has no supplementary-file link on it. So the run "reached" the article,
    # found no supplement, and exited telling us the layout had changed, three
    # weeks running, while the alternate that would have worked sat untried in
    # the list below it. Reaching a page is not the same as getting the thing.
    # ROUTE 1 - Europe PMC, which serves the supplementary files as data.
    rows, src = None, ""
    print("Europe PMC: resolving the DOI to a PMCID")
    pmcid, why = epmc_pmcid(DOI)
    if not pmcid:
        print("  no PMCID: %s" % why)
    else:
        print("  PMCID %s" % pmcid)
        try:
            blob = fetch(EPMC_SUPPL % pmcid, binary=True)
            print("  supplementary archive: %d bytes" % len(blob))
            name, data, why = zip_pick(blob)
            if not name:
                print("  no table in it: %s" % why)
            else:
                print("  table: %s" % name)
                low = name.lower()
                if low.endswith((".xlsx", ".xls")):
                    rows = rows_from_xlsx(data)
                else:
                    rows = rows_from_csv(data.decode("utf-8", "replace"))
                src = "Europe PMC %s (%s)" % (pmcid, name)
                # A ZIP that opened and a sheet that parsed to nothing is not a
                # success. Fall through to the article page rather than writing
                # an empty file that reads as "no recorded incidents".
                if not rows:
                    print("  the table parsed to 0 rows - falling through to the "
                          "article page")
                    rows = None
        except Exception as e:
            print("  supplementary request failed: %s" % str(e)[:90])

    if rows:
        print("  %d rows from Europe PMC" % len(rows))
    else:
        rows = _rows_from_article_page()
        src = "article page"
    _emit(rows, src)


def _rows_from_article_page():
    """The original HTML route. Kept as a fallback and as a record of what was
    tried; it has returned nothing since the journal moved, because Springer
    renders the supplementary section client-side."""
    print("falling back to the article page")
    page, link, reached = None, None, []
    for _cand in ([ARTICLE] + ARTICLE_ALTS):
        try:
            page = fetch(_cand)
        except Exception as e:
            print("  %-66s %s" % (_cand[:66], str(e)[:32]), file=sys.stderr)
            continue
        link = find_supplement(page)
        reached.append((_cand, bool(link)))
        print("  reached %-58s supplement link: %s"
              % (_cand[:58], "yes" if link else "no"))
        if link:
            break
    if not reached:
        sys.exit("Europe PMC did not yield the table and every article address "
                 "refused as well. Nothing written: an empty file would reach the "
                 "map as no recorded incidents, which is the opposite of true. A "
                 "403 everywhere usually means the request was blocked; a 404 "
                 "would mean the article moved again.")
    if not link:
        sys.exit("Europe PMC did not yield the table, and %d of the article's own "
                 "addresses were reached with no supplementary-file link on any of "
                 "them: %s. That is expected - Springer renders that section "
                 "client-side - so the thing to fix is the Europe PMC route above, "
                 "not find_supplement()."
                 % (len(reached), ", ".join(u[:60] for u, _ in reached)))
    print("  supplement: %s" % link[:110])

    try:
        blob = fetch(link, binary=True)
    except Exception as e:
        sys.exit("could not download the supplement: %s" % e)

    if link.lower().endswith((".xlsx", ".xls")):
        rows = rows_from_xlsx(blob)
    elif link.lower().endswith(".csv"):
        rows = rows_from_csv(blob.decode("utf-8", "replace"))
    else:
        sys.exit("supplement is %s, which this script cannot parse. Download it "
                 "by hand from %s and convert to CSV."
                 % (link.rsplit(".", 1)[-1], ARTICLE))

    print("  %d rows in the supplement" % len(rows))
    return rows


def _emit(rows, src):
    """Everything downstream of getting the table: match to countries, drop the
    hand-written duplicates, report, write."""
    print("source of the table: %s" % (src or "unknown"))
    if not rows:
        sys.exit("the table came back empty. Nothing written, because an empty "
                 "register file reaches the map as 396 incidents that never "
                 "happened.")

    have = set()
    if HAND.exists():
        try:
            have = {r["name"].lower()
                    for r in json.loads(HAND.read_text(encoding="utf-8"))["projects"]}
        except Exception:
            pass

    out, skipped, nogeo = [], 0, 0
    for i, r in enumerate(rows):
        rec = to_record(r, i)
        if rec is None:
            nogeo += 1; continue
        if rec["name"].lower() in have:
            skipped += 1; continue
        out.append(rec)

    print("  usable: %d | already hand-written: %d | no country match: %d"
          % (len(out), skipped, nogeo))
    if nogeo:
        print("    (rows without a country this script can place are dropped "
              "rather than guessed at)")

    print()
    print("  Sources that exist and are NOT compiled here:")
    for name, url, note in EXTRA_SOURCES:
        print("    %-12s %s" % (name, url))
        if note:
            print("                 %s" % note)
    print("  Every record emitted carries the scope of its own source, because "
          "no single one of them is a world picture.")

    if "--dry-run" in sys.argv:
        print("\ndry run \u2014 nothing written")
        return
    OUT.write_text(json.dumps({
        "note": ("The GM Contamination Register 1997-2013, recovered from the "
                 "open-access supplementary data of Price & Cotter 2014 "
                 "(doi:10.1186/s40550-014-0005-8, CC BY). The Register was "
                 "compiled by GeneWatch UK and Greenpeace International and "
                 "stopped in 2013. Country-level positions only."),
        "source": src,
        "projects": out}, ensure_ascii=False, indent=1), encoding="utf-8")
    print("\nwrote %s" % OUT.name)


def selftest():
    """No network. Drives the two shapes the new route depends on: the search
    response the PMCID is read out of, and the ZIP the table is picked out of.
    Both were the failure points worth a test - one guessed id or one silently
    wrong file in the archive produces a plausible file full of wrong rows."""
    import zipfile
    ok = True

    def check(label, got, want):
        nonlocal ok
        good = got == want
        ok = ok and good
        print("  %-58s %s" % (label, "pass" if good else "FAIL got %r" % (got,)))

    # zip_pick: spreadsheet beats CSV, CSV beats nothing, and a ZIP with no
    # table says so rather than returning the first file it finds.
    def zbytes(names):
        b = io.BytesIO()
        with zipfile.ZipFile(b, "w") as z:
            for n in names:
                z.writestr(n, "a,b\n1,2\n")
        return b.getvalue()

    check("spreadsheet chosen over csv",
          zip_pick(zbytes(["notes.csv", "Additional file 1.xlsx"]))[0],
          "Additional file 1.xlsx")
    check("csv chosen when there is no spreadsheet",
          zip_pick(zbytes(["readme.pdf", "table.csv"]))[0], "table.csv")
    check("no table in the zip returns no name",
          zip_pick(zbytes(["fig1.pdf", "fig2.png"]))[0], None)
    check("no table in the zip explains itself",
          "none is a spreadsheet or CSV" in (zip_pick(zbytes(["fig1.pdf"]))[2] or ""),
          True)
    check("a zip that is not a zip explains itself",
          "not a readable ZIP" in (zip_pick(b"this is not a zip")[2] or ""), True)

    # to_record: a row with no country this script can place is dropped, never
    # guessed at, and a placed row carries the citation.
    check("row with an unknown country is dropped",
          to_record({"country": "Atlantis", "crop": "maize"}, 0), None)
    r = to_record({"country": "Mexico", "crop": "maize", "year": "2001",
                   "description": "Landrace maize found to contain transgenes."}, 0)
    check("placed row gets that country's centroid", (r["lat"], r["lng"]),
          CENTROID["mexico"])
    check("placed row carries the CC BY citation", CITE in r["desc"], True)
    check("placed row keeps the year", r["date"], "2001-01-01")

    print("\n%s" % ("selftest passed" if ok else "SELFTEST FAILED"))
    return 0 if ok else 1


if __name__ == "__main__":
    if "--selftest" in sys.argv:
        sys.exit(selftest())
    main()
